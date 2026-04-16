"""迁移报告路由 — CRUD + 下载"""
import logging
from io import BytesIO
from flask import Blueprint, jsonify, request, send_file, make_response
from datetime import datetime

from app.services.migration.report_service import ReportService

logger = logging.getLogger(__name__)

report_bp = Blueprint('report', __name__)


@report_bp.route('/tasks/<int:task_id>/report', methods=['POST'])
def generate_report(task_id):
    """从任务执行结果生成迁移报告"""
    try:
        report = ReportService.create_from_task(task_id)
        return jsonify(success=True, data={'report_id': report.id}, message='报告生成完成')
    except ValueError as e:
        return jsonify(success=False, error='NOT_FOUND', message=str(e)), 404
    except Exception as e:
        logger.error(f'生成报告失败: {e}')
        return jsonify(success=False, error='INTERNAL', message=str(e)), 500


# ════════════════════════════════════════════════
# 报告管理 API（持久化）
# ════════════════════════════════════════════════

@report_bp.route('/reports', methods=['GET'])
def list_reports():
    """分页查询报告列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    result = ReportService.list_reports(page=page, page_size=page_size)
    return jsonify(success=True, data=result)


@report_bp.route('/reports/<int:report_id>', methods=['GET'])
def get_report(report_id):
    """获取单个报告详情（含明细）"""
    report = ReportService.get_report(report_id)
    if not report:
        return jsonify(success=False, error='NOT_FOUND', message='报告不存在'), 404
    return jsonify(success=True, data=report)


@report_bp.route('/reports/<int:report_id>', methods=['DELETE'])
def delete_report(report_id):
    """删除报告"""
    ok = ReportService.delete_report(report_id)
    if not ok:
        return jsonify(success=False, error='NOT_FOUND', message='报告不存在'), 404
    return jsonify(success=True, message='报告已删除')


@report_bp.route('/reports/<int:report_id>/download', methods=['GET'])
def download_report(report_id):
    """下载单个报告（Excel 格式）"""
    fmt = request.args.get('format', 'excel')
    report = ReportService.get_report(report_id)
    if not report:
        return jsonify(success=False, error='NOT_FOUND', message='报告不存在'), 404

    if fmt == 'json':
        # JSON 下载
        from flask import Response
        return Response(
            json.dumps(report, ensure_ascii=False, indent=2),
            mimetype='application/json',
            headers={'Content-Disposition': f'attachment;filename=migration-report-{report_id}.json'}
        )
    else:
        # Excel 下载
        return _build_excel_download(report, report_id)


@report_bp.route('/batch-download', methods=['POST'])
def batch_download():
    """批量下载（ZIP 打包 Excel）"""
    data = request.get_json() or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify(success=False, error='NO_IDS', message='请提供报告 ID 列表'), 400

    try:
        import zipfile
        import os

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for rid in ids:
                report = ReportService.get_report(int(rid))
                if not report: continue
                excel_data = _build_excel_bytes(report)
                zf.writestr(f'migration-report-{rid}.xlsx', excel_data)

        zip_buffer.seek(0)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'reports-batch-{ts}.zip',
        )
    except Exception as e:
        logger.error(f'批量下载失败: {e}')
        return jsonify(success=False, error='BATCH_FAILED', message=str(e)), 500


# ── 向后兼容的旧接口 ─────────────────────────────

@report_bp.route('/tasks/<int:task_id>/report', methods=['GET'])
def get_task_report(task_id):
    """获取任务的迁移报告（向后兼容，优先查已有报告）"""
    # 尝试从数据库查找该任务的报告
    from app.extensions import db
    from app.models.report import MigrationReport
    existing = db.session.query(MigrationReport).filter_by(task_id=task_id).first()
    if existing:
        report = ReportService.get_report(existing.id)
        return jsonify(success=True, data=report or {})
    return jsonify(success=True, data={})


@report_bp.route('/tasks/<int:task_id>/report/details', methods=['GET'])
def get_report_details(task_id):
    """获取报告明细（向后兼容）"""
    category = request.args.get('category')
    from app.extensions import db
    from app.models.report import MigrationReport, ReportDetail

    report = db.session.query(MigrationReport).filter_by(task_id=task_id).first()
    if not report:
        return jsonify(success=True, data={'details': []})

    query = db.session.query(ReportDetail).filter_by(report_id=report.id)
    if category:
        query = query.filter_by(category=category)

    details = query.order_by(ReportDetail.created_at).all()
    return jsonify(success=True, data={
        'details': [_detail_to_dict(d) for d in details]
    })


@report_bp.route('/tasks/<int:task_id>/report/export', methods=['GET'])
def export_report(task_id):
    """导出报告（向后兼容）"""
    fmt = request.args.get('format', 'json')
    from app.extensions import db
    from app.models.report import MigrationReport

    report_record = db.session.query(MigrationReport).filter_by(task_id=task_id).first()
    if not report_record:
        return jsonify(success=True, data={}, message=f'无报告数据')

    report = ReportService.get_report(report_record.id)
    if fmt == 'json':
        from flask import Response
        return Response(
            json.dumps(report, ensure_ascii=False, indent=2),
            mimetype='application/json',
            headers={'Content-Disposition': f'attachment;filename=task-{task_id}-report.json'}
        )
    else:
        return _build_excel_download(report, report_record.id)


# ════════════════════════════════════════════════
# Excel 构建辅助
# ════════════════════════════════════════════════

def _build_excel_download(report, report_id):
    """构建 Excel 文件并返回下载响应"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        # 如果没有 openpyxl，回退到 CSV
        return _csv_fallback(report, report_id)

    wb = Workbook()
    ws = wb.active
    ws.title = '迁移摘要'

    # 封面
    ws['A1'] = 'CLB 配置迁移报告'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:B1')
    ws['A3'] = '生成时间'
    ws['B3'] = (report.get('generated_at') or '').replace('T', ' ')
    ws['A4'] = '总项'
    ws['B4'] = report.get('total_items', 0)
    ws['A5'] = '成功'
    ws['B5'] = report.get('success_count', 0)
    ws['A6'] = '失败'
    ws['B6'] = report.get('failed_count', 0)
    ws['A7'] = '跳过'
    ws['B7'] = report.get('skipped_count', 0)
    ws['A8'] = '结论'
    ws['B8'] = report.get('report_summary', '')

    # 明细 sheet
    ws_detail = wb.create_sheet(title='操作明细')
    headers = ['序号', '类型', '描述', '状态', '耗时(ms)', '错误信息']
    for col, h in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    for idx, detail in enumerate(report.get('details', []), 2):
        ws_detail.cell(row=idx, column=1, value=idx - 1)
        ws_detail.cell(row=idx, column=2, value=detail.get('operation_type', ''))
        ws_detail.cell(row=idx, column=3, value=detail.get('operation_desc', ''))
        ws_detail.cell(row=idx, column=4, value=detail.get('category', ''))
        ws_detail.cell(row=idx, column=5, value=detail.get('duration_ms') or '')
        ws_detail.cell(row=idx, column=6, value=detail.get('error_message') or detail.get('incompatible_reason') or '')

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'migration-report-{report_id}.xlsx')


def _build_excel_bytes(report):
    """构建 Excel 返回字节流（用于 ZIP 打包）"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = '迁移摘要'
    ws['A1'] = 'CLB 配置迁移报告'
    ws.merge_cells('A1:B1')
    ws['A3'] = '总项'; ws['B3'] = report.get('total_items', 0)
    ws['A4'] = '成功'; ws['B4'] = report.get('success_count', 0)
    ws['A5'] = '失败'; ws['B5'] = report.get('failed_count', 0)
    ws['A6'] = '跳过'; ws['B6'] = report.get('skipped_count', 0)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _csv_fallback(report, report_id):
    """openpyxl 未安装时回退到 CSV"""
    import csv
    from flask import Response
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(['序号', '类型', '描述', '状态', '错误信息'])
    for idx, d in enumerate(report.get('details', []), 1):
        writer.writerow([idx, d.get('operation_type'), d.get('operation_desc'),
                        d.get('category'), d.get('error_message')])
    return Response(buf.getvalue(), mimetype='text/csv',
                    headers={'Content-Disposition': f'attachment; filename=report-{report_id}.csv'})


def _detail_to_dict(d):
    return {
        'id': d.id,
        'category': d.category,
        'operation_type': d.operation_type,
        'operation_desc': d.operation_desc,
        'source_config': d.source_config,
        'target_config': d.target_config,
        'error_message': d.error_message,
        'incompatible_reason': d.incompatible_reason,
        'duration_ms': d.duration_ms,
    }
